from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import shutil
import re
import io
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from functools import wraps

app = Flask(__name__)
CORS(app)

BASE_DIR = "data"
UPLOAD_DIR = "uploads"

os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ---------------- BASIC ----------------
@app.route("/")
def home():
    return "Backend running"

def safe_name(name):
    return name.replace(" ", "_")

def make_shift_id(exam_date, exam_time):
    date_part = exam_date.replace("/", "-")
    time_part = exam_time.replace(":", "-")
    time_part = re.sub(r"\s+", "", time_part)
    return f"{date_part}_{time_part}"

@app.route("/exams")
def public_exams():
    exams = []
    for folder in os.listdir(BASE_DIR):
        exam_dir = os.path.join(BASE_DIR, folder)
        scheme_path = os.path.join(exam_dir, "marking_scheme.xlsx")
        if not os.path.exists(scheme_path):
            continue

        wb = load_workbook(scheme_path)
        ws = wb.active
        scheme = {r[0]: r[1] for r in ws.iter_rows(min_row=2, values_only=True)}

        exams.append({
            "exam_name": scheme.get("Exam Name")
        })

    return jsonify(exams)


# ---------------- ADMIN: LIST EXAMS ----------------
@app.route("/admin/exams", methods=["GET"])
def admin_list_exams():
    exams = []
    for folder in os.listdir(BASE_DIR):
        exam_dir = os.path.join(BASE_DIR, folder)
        scheme_path = os.path.join(exam_dir, "marking_scheme.xlsx")
        if not os.path.exists(scheme_path):
            continue

        wb = load_workbook(scheme_path)
        ws = wb.active
        scheme = {r[0]: r[1] for r in ws.iter_rows(min_row=2, values_only=True)}

        exams.append({
            "exam_name": scheme.get("Exam Name"),
            "correct": scheme.get("Correct"),
            "wrong": scheme.get("Wrong"),
            "na": scheme.get("NA")
        })
    return jsonify(exams)

# ---------------- ADMIN: CREATE EXAM ----------------
@app.route("/admin/create-exam", methods=["POST"])
def create_exam():
    data = request.json
    exam_name = data.get("exam_name")
    subjects = data.get("subjects", [])
    correct = data.get("correct")
    wrong = data.get("wrong")
    na = data.get("na")

    if not exam_name or not subjects:
        return jsonify({"error": "Invalid data"}), 400

    exam_dir = os.path.join(BASE_DIR, safe_name(exam_name))
    os.makedirs(exam_dir, exist_ok=True)

    # marking scheme
    wb = Workbook()
    ws = wb.active
    ws.append(["Key", "Value"])
    ws.append(["Exam Name", exam_name])
    ws.append(["Correct", correct])
    ws.append(["Wrong", wrong])
    ws.append(["NA", na])
    wb.save(os.path.join(exam_dir, "marking_scheme.xlsx"))

    # subjects
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Subject Name", "Max Marks", "Count In Total"])
    for s in subjects:
        ws2.append([s["name"], s["max_marks"], "YES" if s["count_in_total"] else "NO"])
    wb2.save(os.path.join(exam_dir, "subjects.xlsx"))

    # responses template
    wb3 = Workbook()
    ws3 = wb3.active
    header = ["Name", "Roll", "Category", "Gender", "State", "Final Marks"]
    for s in subjects:
        base = s["name"]
        header += [
            f"{base}_Attempt",
            f"{base}_R",
            f"{base}_W",
            f"{base}_NA",
            f"{base}_Marks"
        ]
    ws3.append(header)
    wb3.save(os.path.join(exam_dir, "responses.xlsx"))

    return jsonify({"status": "success"})

# ---------------- ADMIN: DELETE EXAM ----------------
@app.route("/admin/delete-exam", methods=["POST"])
def delete_exam():
    exam = request.json.get("exam_name")
    path = os.path.join(BASE_DIR, safe_name(exam))
    if not os.path.exists(path):
        return jsonify({"error": "Exam not found"}), 404
    shutil.rmtree(path)
    return jsonify({"status": "deleted"})

# ---------------- READ HELPERS ----------------
def read_marking_scheme(exam):
    wb = load_workbook(os.path.join(BASE_DIR, safe_name(exam), "marking_scheme.xlsx"))
    return {r[0]: r[1] for r in wb.active.iter_rows(min_row=2, values_only=True)}

def read_subjects(exam):
    wb = load_workbook(os.path.join(BASE_DIR, safe_name(exam), "subjects.xlsx"))
    return [{
        "name": r[0],
        "count_in_total": r[2] == "YES"
    } for r in wb.active.iter_rows(min_row=2, values_only=True)]

# ---------------- PARSE RESPONSE HTML ----------------
def parse_response_html(html, scheme, subjects):
    soup = BeautifulSoup(html, "lxml")
    section_map = {}
    order = []
    current = None

    for el in soup.find_all("div"):
        if el.get("class") == ["section-lbl"]:
            current = el.get_text(strip=True)
            section_map[current] = {"c": 0, "w": 0, "n": 0}
            order.append(current)

        if el.get("class") == ["question-pnl"] and current:
            chosen = None
            correct = None

            ch = el.find("td", string=re.compile("Chosen Option"))
            if ch:
                v = ch.find_next_sibling("td").get_text(strip=True)
                if v != "--":
                    chosen = int(v)

            rt = el.find("td", class_="rightAns")
            if rt:
                m = re.search(r"(\d)\.", rt.get_text())
                if m:
                    correct = int(m.group(1))

            if chosen is None:
                section_map[current]["n"] += 1
            elif chosen == correct:
                section_map[current]["c"] += 1
            else:
                section_map[current]["w"] += 1

    results = []
    final = 0

    for i, sec in enumerate(order):
        c = section_map[sec]["c"]
        w = section_map[sec]["w"]
        n = section_map[sec]["n"]

        marks = c * scheme["Correct"] + w * scheme["Wrong"] + n * scheme["NA"]
        sub = subjects[i]

        results.append({
            "name": sub["name"],
            "attempt": c + w,
            "r": c,
            "w": w,
            "na": n,
            "marks": marks,
            "count_in_total": sub["count_in_total"]
        })

        if sub["count_in_total"]:
            final += marks

    return final, results

# ---------------- SAVE RESULT ----------------
def save_user_result(exam_name, shift_id, base_data, subject_results):
    exam_dir = os.path.join(BASE_DIR, safe_name(exam_name))
    path = os.path.join(exam_dir, f"responses_{shift_id}.xlsx")

    if not os.path.exists(path):
        shutil.copy(os.path.join(exam_dir, "responses.xlsx"), path)

    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    roll_value = base_data[col["Roll"]]
    updated = False

    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row_idx, col["Roll"] + 1).value) == str(roll_value):
            for key in ["Name", "Category", "Gender", "State", "Final Marks"]:
                ws.cell(row_idx, col[key] + 1, base_data[col[key]])
            updated = True
            break

    if not updated:
        row = base_data[:]
        for sub in subject_results:
            row.extend([sub["attempt"], sub["r"], sub["w"], sub["na"], sub["marks"]])
        ws.append(row)

    wb.save(path)
    calculate_shift_ranks(path)

# ---------------- EVALUATE (HTML UPLOAD – OPTION 1) ----------------
@app.route("/evaluate", methods=["POST"])
def evaluate():
    exam = request.form.get("exam_name")
    category = request.form.get("category")
    gender = request.form.get("gender")
    state = request.form.get("state")
    file = request.files.get("file")

    if not all([exam, category, gender, state, file]):
        return jsonify({"error": "Missing fields"}), 400

    html = file.read().decode("utf-8", errors="ignore")
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text(" ")

    date_match = re.search(r"\d{2}/\d{2}/\d{4}", text)
    time_match = re.search(
        r"\d{1,2}:\d{2}\s*(AM|PM)\s*-\s*\d{1,2}:\d{2}\s*(AM|PM)", text
    )

    if not date_match or not time_match:
        return jsonify({"error": "Unable to detect exam date or shift"}), 400

    shift = make_shift_id(date_match.group(), time_match.group())

    def extract(label):
        td = soup.find("td", string=re.compile(label))
        return td.find_next_sibling("td").get_text(strip=True) if td else None

    name = extract("Candidate Name")
    roll = extract("Roll")

    if not name or not roll:
        return jsonify({"error": "Candidate details not found"}), 400

    scheme = read_marking_scheme(exam)
    subjects = read_subjects(exam)

    final, subject_results = parse_response_html(html, scheme, subjects)

    save_user_result(
        exam,
        shift,
        [name, roll, category, gender, state, final],
        subject_results
    )

    return jsonify({"status": "saved", "roll": roll})

# ---------------- RESULT API ----------------
@app.route("/result")
def get_result():
    exam = request.args.get("exam")
    roll = request.args.get("roll")

    exam_dir = os.path.join(BASE_DIR, safe_name(exam))
    subjects_cfg = read_subjects(exam)

    for file in os.listdir(exam_dir):
        if not file.startswith("responses_"):
            continue

        path = os.path.join(exam_dir, file)
        wb = load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        for r in ws.iter_rows(min_row=2, values_only=True):
            if str(r[col["Roll"]]) == str(roll):
                shift_id = file.replace("responses_", "").replace(".xlsx", "")
                counted, qualifying, final = [], [], 0

                for sub in subjects_cfg:
                    base = sub["name"]
                    item = {
                        "name": base,
                        "attempt": r[col[f"{base}_Attempt"]],
                        "na": r[col[f"{base}_NA"]],
                        "r": r[col[f"{base}_R"]],
                        "w": r[col[f"{base}_W"]],
                        "marks": r[col[f"{base}_Marks"]]
                    }
                    if sub["count_in_total"]:
                        counted.append(item)
                        final += item["marks"]
                    else:
                        qualifying.append(item)

                return jsonify({
                    "exam": exam,
                    "shift_id": shift_id,
                    "candidate": {
                        "name": r[col["Name"]],
                        "roll": r[col["Roll"]],
                        "category": r[col["Category"]],
                        "gender": r[col["Gender"]],
                        "state": r[col["State"]],
                        "rank": r[col.get("Rank")]
                    },
                    "total_candidates": ws.max_row - 1,
                    "counted_subjects": counted,
                    "qualifying_subjects": qualifying,
                    "final_marks": final
                })

    return jsonify({"error": "Candidate not found"}), 404

# ---------------- RANK ----------------
def calculate_shift_ranks(path):
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    if "Rank" not in col:
        ws.cell(1, len(headers) + 1, "Rank")
        col["Rank"] = len(headers)

    rows = [(r, ws.cell(r, col["Final Marks"] + 1).value)
            for r in range(2, ws.max_row + 1)]
    rows.sort(key=lambda x: x[1], reverse=True)

    rank = 1
    prev = None
    count = 0
    for r, m in rows:
        if m != prev:
            rank = count + 1
        ws.cell(r, col["Rank"] + 1, rank)
        prev = m
        count += 1

    wb.save(path)

def fetch_result_data(exam, roll):
    # EVERYTHING you currently do inside /result
    # XLSX loading
    # roll matching
    # marks calculation
    # return dict

    if not exam or not roll:
        return {"error": "Missing exam or roll"}

    # example return (structure must match /result)
    return {
        "exam": exam.replace("_", " "),
        "candidate": {
            "name": "ABC",
            "roll": roll,
            "category": "UR",
            "state": "Delhi",
            "rank": 123
        },
        "final_marks": 245.5,
        "counted_subjects": [],
        "qualifying_subjects": []
    }

@app.route("/result-pdf")
def result_pdf():
    exam = request.args.get("exam")
    roll = request.args.get("roll")

    result = fetch_result_data(exam, roll)
    if "error" in result:
        return jsonify(result), 400

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # ---------- WATERMARK ----------
    c.saveState()
    c.setFont("Helvetica-Bold", 60)
    c.setFillGray(0.9)
    c.translate(width / 2, height / 2)
    c.rotate(45)
    c.drawCentredString(0, 0, "RankPred")
    c.restoreState()

    y = height - 50

    # ---------- TITLE ----------
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, y, result["exam"])
    y -= 30

    # ---------- CANDIDATE INFO ----------
    c.setFont("Helvetica", 11)
    cand = result["candidate"]

    c.drawString(50, y, f"Name: {cand['name']}"); y -= 15
    c.drawString(50, y, f"Roll: {cand['roll']}"); y -= 15
    c.drawString(50, y, f"Category: {cand['category']}"); y -= 15
    c.drawString(50, y, f"State: {cand['state']}"); y -= 20

    # ---------- MARKS ----------
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, f"Final Marks: {result['final_marks']}")
    y -= 25

    # ---------- SUBJECTS ----------
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, y, "Subject-wise Marks")
    y -= 15

    c.setFont("Helvetica", 10)
    for sub in result["counted_subjects"]:
        c.drawString(
            50, y,
            f"{sub['name']} | Marks: {sub['marks']} | R: {sub['r']} | W: {sub['w']}"
        )
        y -= 12

        if y < 50:
            c.showPage()
            y = height - 50

    c.showPage()
    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{roll}_result.pdf",
        mimetype="application/pdf"
    )

# ---------------- START ----------------
if __name__ == "__main__":
    app.run()
